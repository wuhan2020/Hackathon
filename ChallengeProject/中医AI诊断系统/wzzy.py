# coding=utf-8
import streamlit as st
import numpy as np
import pandas as pd


qzyz = 0.2    #权重因子0-1，就是在一串症状中，有多少比例的症状是有效的

st.title("中医AI问诊V1.1")

import os
path_list = os.listdir()
path_name=[]
for i in path_list:
	if i.split(".")[1] == 'xlsx':
		path_name.append(i.split(".")[0])
zhuanjiaXZ=st.selectbox('请选择一个专家AI:',(path_name))
zhuangjiaXZ_file=str(zhuanjiaXZ)+'.xlsx'
from openpyxl import load_workbook
wb = load_workbook(zhuangjiaXZ_file)
Sheet1 = wb["Sheet1"]
zhenzhuang=set()
st.write(Sheet1.cell(1,1).value)
qzyz = Sheet1.cell(1,12).value     #从表中读取权重因子
if qzyz <0 or qzyz>1 :
	qzyz =0.2
for i in range(3,Sheet1.max_row+1):
	for j in range(1,11):
		zhenzhuang.add(Sheet1.cell(i,j).value)
zhenzhuang.discard(None)
zhenzhuanglist=list(zhenzhuang)
zhenzhuanglist.sort()
zhenzhuangXZ = st.multiselect(
	'选择你的症状（多选）：',(zhenzhuanglist))
zhenzhuangXZ_set=set(zhenzhuangXZ)



#开始症状分析
zhenzhuangGL=set()
yaofang=set()
zhenduan=set()
quanzhong=[0,0,0] #症状对应每个方剂的权重
for i in range(3,Sheet1.max_row+1):
	for j in range(1,11):
		zhenzhuangGL.add(Sheet1.cell(i,j).value)
	zhenzhuangGL.discard(None)
	z = zhenzhuangGL.intersection(zhenzhuangXZ_set)   #两个集合的交集
	if len(zhenzhuangGL)==0 :
		quanzhong.append(len(z))
	else:
		quanzhong.append(len(z)/len(zhenzhuangGL))
	zhenzhuangGL.clear()


import copy 
t = copy.deepcopy(quanzhong)
# 求m个最大的数值及其索引
max_number = list()
max_index = list()
for _ in range(3):
    number = max(t)
    index = t.index(number)
    t[index] = 0
    max_number.append(number)
    max_index.append(index)
t = []

if max_number[2] < qzyz :  #移除权重过低的答案
	max_number.pop()
	max_index.pop()
if max_number[1] < qzyz :  
	max_number.pop()
	max_index.pop()
if max_number[0] < qzyz :  	
	st.write('未能做出诊断，请重新选择症状')
else:
	for i in range(len(max_index)) :
		zhenduan.add(Sheet1.cell(max_index[i],11).value)   #最接近症状的诊断
		zhenduan.add(Sheet1.cell(max_index[i],12).value)
		for j in range(16,30):
			yaofang.add(Sheet1.cell(max_index[i],j).value)
	zhenduan.discard(None)
	yaofang.discard(None)
	zhenduan_list=list(zhenduan)
	zhenduan_list.sort()
	yaofang_list=list(yaofang)
	yaofang_list.sort()
	str = " "
	st.write('诊断及建议: ')
	st.write(str.join(zhenduan_list))
	st.write('建议中药：',str.join(yaofang_list))
	st.write('中药剂量及服用有严格的要求，请在合格中医师指导下进行，我们不对服用中药后引起的任何后果负责')


#© 2020 GitHub, Inc.